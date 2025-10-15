import pygame
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from scripts.map import Map
from scripts.controller import Controller
from scripts.extra import ExtraFunc

pygame.init()

extra_func = ExtraFunc()

button_play_AI_path = r"assets\images\ButtonPlay.png"
button_select_algorithm_path = r"assets\images\ButtonAlgorithm.png"
button_extract_path = r"assets\images\buttonExtract.png"

sound_button_clicked_path = r"D:\tam\stu\triTueNhanTao\gameSokoban\sounds\buttonClicked.mp3"

class Gameplay():
    def __init__(self, window):
        self.window = window
        self.screen = self.window.screen
        self.mode = self.window.get_data("mode")
        self.level = self.window.get_data("level")
        self.data_human = self.window.get_data("game_play_human").copy()
        self.data_AI = self.window.get_data("game_play_AI").copy()
        self.map = Map(self.window)
        self.map.create_map()
        self.controller = self.data_AI["Controller"]
        if self.controller is None:
            self.controller = Controller(self.window)
        self.text_esc = None
        self.text_clock_human = None
        self.text_status_human = None
        self.text_status_AI = None
        self.button_play_AI = None
        self.button_select_algorithm = None
        self.button_extract = None
        if self.mode == "human":
            self.text_clock_human = TextClockHuman(self.window, 30, self.data_human["TextClockHuman"].copy())
            self.text_status_human = TextStatusHuman(self.window, self.data_human["TextStatusHuman"].copy())
            self.text_esc = TextEsc(self.window, self.data_human["TextEsc"].copy())
        if self.mode == "AI":
            self.button_play_AI = ButtonPlayAI(self.window, self.data_AI["ButtonPlayAI"].copy())
            self.text_status_AI = TextStatusAI(self.window, self.data_AI["TextStatusAI"].copy())
            self.text_esc = TextEsc(self.window, self.data_AI["TextEsc"].copy())
            self.button_select_algorithm = ButtonSelectAlgorithm(self.window, self.data_AI["ButtonSelectAlgorithm"].copy())
            self.button_extract = ButtonExtract(self.window)
        self.clock = pygame.time.Clock()
        self.fps = self.window.get_data("fps")

    def handle_events(self, event):
        if self.mode == "human":
            if event and event.type == pygame.KEYDOWN:
                if not self.controller.check_win(self.mode) and not self.text_clock_human.check_time_out():
                    self.controller.handle_human_action(event.key)
                    self.map.create_map()
            if not event:
                if not self.controller.check_win(self.mode) and self.text_clock_human.check_time_out():
                    self.text_status_human.change_text("Bạn thua!")
                    self.text_clock_human.stop_clock()
                elif self.check_win():
                    self.text_status_human.change_text("Bạn thắng!")
                    self.text_clock_human.stop_clock()
        if self.mode == "AI":
            if not event:
                if self.text_status_AI and self.text_status_AI.text == "Đang mở hộp thoại thuật toán!":
                    if self.button_select_algorithm and self.button_select_algorithm.check_goal_zoom():
                        self.window.set_data("status_screen", "algorithmSelector")
                        self.data_AI["Controller"] = self.controller
                        self.data_AI["ButtonPlayAI"] = self.button_play_AI.get_data()
                        self.window.set_data("game_play_AI", self.data_AI)
                if self.text_status_AI and self.text_status_AI.text == "AI đang suy nghĩ...":
                    if self.controller:
                            self.controller.run_algorithm()
                    self.button_play_AI.clicked()
                    self.text_esc.show = True
                    if self.check_win():
                        self.text_status_AI.change_text("AI đã tìm ra lời giải!")
                    else:
                        self.text_status_AI.change_text("AI không tìm ra được lời giải trong điều kiện giới hạn!")
                if self.controller.handle_AI_action():
                    self.map.create_map()
            if event and event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1:
                    pos = pygame.mouse.get_pos()
                    if self.button_play_AI and self.button_play_AI.check_collided(pos):
                        self.button_play_AI.run_music()
                        if self.text_status_AI and self.controller:
                            if self.controller.check_algorithm_selecting() and not self.button_play_AI.check_clicked():
                                self.text_status_AI.change_text("AI đang suy nghĩ...")
                                self.text_esc.show = False
                            elif self.button_play_AI.check_clicked():
                                self.text_status_AI.change_text("Hãy chọn thuật toán khác!")
                            else:
                                self.text_status_AI.change_text("Chưa chọn thuật toán!")
                    if self.button_select_algorithm and self.button_select_algorithm.check_collided(pos):
                        self.button_select_algorithm.run_music()
                        if self.text_status_AI and not self.check_complete_AI():
                            self.text_status_AI.change_text("Hãy đợi mô phỏng thuật toán hoàn thành!")
                        elif self.text_status_AI:
                            self.text_status_AI.change_text("Đang mở hộp thoại thuật toán!")
                            self.button_select_algorithm.clicked()
                    if self.button_extract and self.button_extract.check_collided(pos):
                        self.button_extract.run_music()
                        name_algorithm = self.window.get_data("algorithm")
                        if len(name_algorithm) > 0 and self.text_status_AI:
                            if self.controller and self.controller.algorithm_detail_dict:
                                self.button_extract.extract_file(self.controller.algorithm_detail_dict, name_algorithm)
                                self.text_status_AI.change_text("Đã xuất file!")
                            else:
                                self.text_status_AI.change_text("Không có dữ liệu để xuất!")
                        else:
                            self.text_status_AI.change_text("Chưa chọn thuật toán!")
                            

        if event and event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                if self.check_complete_AI():
                    self.set_data()
                    self.window.set_data("status_screen", "menu")
                    self.window.set_data("menu_back", True)
                elif self.text_status_AI:
                    self.text_status_AI.change_text("Hãy đợi mô phỏng thuật toán hoàn thành!")
        return True
    
    def set_data(self):
        if self.mode == "human":
            self.data_human["TextClockHuman"] = self.text_clock_human.get_data()
            print(self.data_human["TextClockHuman"])
            self.data_human["TextStatusHuman"] = self.text_status_human.get_data()
            self.data_human["TextEsc"] = self.text_esc.get_data()
            self.window.set_data("game_play_human", self.data_human)
        if self.mode == "AI":
            self.data_AI["Controller"] = self.controller
            self.data_AI["ButtonPlayAI"] = self.button_play_AI.get_data()
            self.data_AI["TextStatusAI"] = self.text_status_AI.get_data()
            self.data_AI["TextEsc"] = self.text_esc.get_data()
            self.data_AI["ButtonSelectAlogorithm"] = self.button_select_algorithm.get_data()
            self.window.set_data("game_play_AI", self.data_AI)

    def draw_ui(self):
        self.screen.fill((0, 0, 0))
        t = time.perf_counter()

        text_list = [
            self.text_esc,
            self.text_clock_human,
            self.text_status_human,
            self.text_status_AI
        ]

        button_list = [
            self.button_play_AI,
            self.button_select_algorithm,
            self.button_extract
        ]

        for button in button_list:
            if button:
                button.effect_movement()

        for text in text_list:
            if text:
                text.draw(t)

        self.map.draw_map()

    def run(self):
        self.draw_ui()
        pygame.display.flip()
        self.clock.tick(self.fps)

    def check_win(self):
        if self.mode == "human":
            return self.controller.check_win(self.mode) and not self.text_clock_human.check_time_out()
        elif self.mode == "AI":
            return self.controller.check_win(self.mode)

    def check_complete_AI(self):
        return self.controller.check_complete_AI()

    def release(self):
        pass

class TextEsc():
    def __init__(self, window, data):
        self.window = window
        self.screen = self.window.screen
        self.font = pygame.font.SysFont("Arial", 15)
        self.text = "Nhấn ESC để quay lại..."
        self.color = (255, 255, 255)
        self.image = self.font.render(self.text, True, self.color)
        pos_center_x = round(self.window.get_data("screen_size")[0] / 2)
        self.pos_center = (pos_center_x, 20)
        self.data = data
        self.show = self.data["show"]
        if self.show is None:
            self.show = True
        self.rect = self.image.get_rect(center=self.pos_center)

    def get_data(self):
        self.data["show"] = self.show
        return self.data

    def draw(self, t, speed_time=0.5):
        alpha = int((abs((t % 1) - speed_time) * 2) * 255)
        self.image.set_alpha(alpha)
        if self.show:
            self.screen.blit(self.image, self.rect)

class TextClockHuman():
    def __init__(self, window, count_time, data):
        self.window = window
        self.screen = self.window.screen
        self.font = pygame.font.SysFont("Arial", 22, True)
        self.count_time = count_time + 1
        self.data = data
        self.remaining_time = None
        self.stop = self.data["stop"]
        if self.stop is None:
            self.stop = False
        self.text = str(self.remaining_time)
        self.start_time = time.perf_counter()
        self.elapsed = self.data["elapsed"]
        if self.elapsed:
            self.start_time = None
        self.color = (255, 255, 255)
        self.image = self.font.render(self.text, True, self.color)
        pos_center_x = round(self.window.get_data("screen_size")[0] / 2)
        self.pos_center = (pos_center_x, 60)
        self.rect = self.image.get_rect(center=self.pos_center)

    def draw(self, t):
        if not self.stop:
            if self.start_time is None:
                self.start_time = t - self.elapsed
            self.elapsed = t - self.start_time
            
        self.remaining_time = max(0, int(self.count_time - self.elapsed))
        mins = self.remaining_time // 60
        secs = self.remaining_time % 60
        text = f"{mins:02d}:{secs:02d}"

        image = self.font.render(text, True, self.color)
        rect = image.get_rect(center=self.pos_center)
        self.screen.blit(image, rect)

    def get_data(self):
        self.data["elapsed"] = self.elapsed
        self.data["stop"] = self.stop
        return self.data

    def stop_clock(self):
        self.stop = True

    def check_time_out(self):
        if self.remaining_time == 0:
            return True
        return False
    
class TextStatusHuman():
    def __init__(self, window, data):
        self.window = window
        self.screen = self.window.screen
        self.font = pygame.font.SysFont("Arial", 30, True)
        self.color = (255, 255, 255)
        self.text_ori = "Bắt đầu!"
        self.data = data
        self.text = self.data["text"]
        if self.text == "":
            self.text = self.text_ori
        self.image = self.font.render(self.text, True, self.color)
        pos_center_x = round(self.window.get_data("screen_size")[0] / 2)
        self.pos_center = (pos_center_x, 120)
        self.rect = self.image.get_rect(center=self.pos_center)

    def get_data(self):
        self.data["text"] = self.text
        return self.data

    def draw(self, t, speed_time=0.5):
        alpha = int((abs((t % 1) - speed_time) * 2) * 255)
        self.image.set_alpha(alpha)
        self.screen.blit(self.image, self.rect)

    def change_text(self, text=None):
        if text:
            self.text = text
        else:
            self.text = self.text_ori
        self.image = self.font.render(self.text, True, self.color)
        self.rect = self.image.get_rect(center=self.pos_center)

class TextStatusAI():
    def __init__(self, window, data):
        self.window = window
        self.screen = self.window.screen
        self.font = pygame.font.SysFont("Arial", 15)
        self.color = (255, 255, 255)
        self.algorithm = self.window.get_data("algorithm")
        self.text_ori = "Hãy chọn thuật toán!"
        if len(self.algorithm) > 0:
            self.text_ori = f"Đã chọn {self.algorithm}!"
        self.data = data
        self.text = self.data["text"]
        if self.text == "":
            self.text = self.text_ori
        self.image = self.font.render(self.text, True, self.color)
        pos_center_x = round(self.window.get_data("screen_size")[0] / 2)
        self.pos_center = (pos_center_x, 220)
        self.rect = self.image.get_rect(center=self.pos_center)

    def get_data(self):
        self.data["text"] = self.text
        return self.data

    def draw(self, t, speed_time=0.5):
        alpha = int((abs((t % 1) - speed_time) * 2) * 255)
        if self.text == "AI đang suy nghĩ...":
            alpha = 255
        self.image.set_alpha(alpha)
        self.screen.blit(self.image, self.rect)

    def change_text(self, text=None):
        if text:
            self.text = text
        else:
            self.text = self.text_ori
        self.image = self.font.render(self.text, True, self.color)
        self.rect = self.image.get_rect(center=self.pos_center)

class ButtonPlayAI():
    def __init__(self, window, data):
        self.window = window
        self.screen = self.window.screen
        self.data = data
        self.image_ori = pygame.image.load(button_play_AI_path)
        self.size = (150, 150)
        self.image_ori = pygame.transform.scale(self.image_ori, self.size)
        self.image = self.image_ori
        pos_init_center_x = round(self.window.get_data("screen_size")[0] / 2)
        pos_init_center_y = round(self.window.get_data("screen_size")[1] / 6)
        self.pos_init_center = (pos_init_center_x, pos_init_center_y)
        self.angle = self.data["angle"]
        self.music = pygame.mixer.Sound(sound_button_clicked_path)
        if self.angle is None:
            self.angle = 0
        self.goal_angle = -90
        self.speed = self.data["speed"]
        if self.speed is None:
            self.speed = 0
        self.rect = self.image.get_rect(center=self.pos_init_center)

    def run_music(self):
        self.music.play()

    def effect_movement(self):
        if self.angle > self.goal_angle:
            self.angle += self.speed
        self.image = pygame.transform.rotate(self.image_ori, self.angle)
        self.rect = self.image.get_rect(center=self.rect.center)

        self.screen.blit(self.image, self.rect)

    def get_data(self):
        self.data["angle"] = self.angle
        self.data["speed"] = self.speed
        return self.data

    def check_collided(self, pos):
        return self.rect.collidepoint(pos)

    def check_clicked(self):
        if self.angle == self.goal_angle:
            return True
        return False

    def clicked(self, speed=-10):
        if self.speed == 0:
            self.speed = speed

class ButtonSelectAlgorithm():
    def __init__(self, window, data):
        self.window = window
        self.screen = self.window.screen
        self.data = data
        self.image_ori = pygame.image.load(button_select_algorithm_path)
        self.size = (120, 120)
        self.image_ori = pygame.transform.scale(self.image_ori, self.size)
        self.image = self.image_ori
        pos_init_center_x = round(self.window.get_data("screen_size")[0] / 4)
        pos_init_center_y = round(self.window.get_data("screen_size")[1] / 6)
        self.music = pygame.mixer.Sound(sound_button_clicked_path)
        self.zoom_init_rate = self.data["zoom_init_rate"]
        self.zoom_goal_rate = 1.0
        self.zoom_current_rate = self.data["zoom_current_rate"]
        if self.zoom_init_rate is None or self.zoom_init_rate != self.zoom_goal_rate:
            self.zoom_init_rate = 0.8
        elif self.zoom_init_rate == self.zoom_goal_rate:
            self.zoom_goal_rate = 1.0

        if self.zoom_current_rate is None:
            self.zoom_current_rate = self.zoom_goal_rate

        self.zoom_current_rate = self.zoom_goal_rate
        self.pos_init_center = (pos_init_center_x, pos_init_center_y)
        self.pos_current_center = self.pos_init_center
        self.rect = self.image.get_rect(center=self.pos_current_center)

    def run_music(self):
        self.music.play()

    def effect_movement(self, speed_zoom=0.01):
        distance_zoom = round(abs(self.zoom_goal_rate - self.zoom_init_rate), 5)
        speed_zoom = extra_func.normalize_speed(distance_zoom, speed_zoom)
        dz = round(self.zoom_goal_rate - self.zoom_current_rate, 5)
        d = extra_func.direction(dz)
        self.zoom_current_rate += d*speed_zoom

        self.image = pygame.transform.scale(self.image_ori, (int(self.size[0]*self.zoom_current_rate), int(self.size[1]*self.zoom_current_rate)))
        self.rect = self.image.get_rect(center=self.rect.center)

        self.screen.blit(self.image, self.rect)

    def check_goal_zoom(self):
        if round(self.zoom_goal_rate - self.zoom_current_rate, 3) == 0:
            print(round(self.zoom_goal_rate - self.zoom_current_rate, 3))
            return True
        print(round(self.zoom_goal_rate - self.zoom_current_rate, 3))
        return False

    def check_collided(self, pos):
        return self.rect.collidepoint(pos)

    def get_data(self):
        self.data["zoom_init_rate"] = self.zoom_init_rate
        self.data['zoom_current_rate'] = self.zoom_current_rate
        return self.data

    def clicked(self):
        self.zoom_init_rate, self.zoom_goal_rate = self.zoom_goal_rate, self.zoom_init_rate

class ButtonExtract():
    def __init__(self, window):
        self.window = window
        self.screen = self.window.screen
        self.image_ori = pygame.image.load(button_extract_path)
        self.size = (100, 100)
        self.image_ori = pygame.transform.scale(self.image_ori, self.size)
        self.image = self.image_ori
        self.screen_size = self.window.get_data("screen_size")
        self.pos_init_centerx = int(3 * self.screen_size[0] // 4)
        self.pos_init_centery = int(1*self.screen_size[1] // 6)
        self.pos_init_center = (self.pos_init_centerx, self.pos_init_centery)
        self.rect = self.image.get_rect(center=self.pos_init_center)
        self.music = pygame.mixer.Sound(sound_button_clicked_path)

    def run_music(self):
        self.music.play()

    def _process_and_write_sheet(self, wb, data_list, sheet_name):
        safe_sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_sheet_name)

        all_keys = set()
        for item in data_list:
            all_keys.update(item.keys())
        header = sorted(list(all_keys))

        # Ghi header vào hàng đầu tiên
        ws.append(header)
        
        # Style cho hyperlink
        hyperlink_font = Font(color="0000FF", underline="single")

        # 3. Ghi dữ liệu vào từng hàng
        for row_idx, item_dict in enumerate(data_list, start=2): # Bắt đầu từ hàng 2
            for col_idx, key in enumerate(header, start=1):
                value = item_dict.get(key)
                cell = ws.cell(row=row_idx, column=col_idx)

                if isinstance(value, dict):
                    # Nếu giá trị là một dict -> tạo sheet mới
                    new_sheet_name = f"{safe_sheet_name}_{key}"
                    self._process_and_write_sheet(wb, [value], new_sheet_name) # [value] vì hàm yêu cầu list
                    
                    # Tạo hyperlink đến sheet mới
                    cell.value = f"[Xem sheet: {new_sheet_name[:31]}]"
                    cell.hyperlink = f"#'{new_sheet_name[:31]}'!A1"
                    cell.font = hyperlink_font

                elif isinstance(value, list) and value and all(isinstance(i, dict) for i in value):
                    new_sheet_name = f"{safe_sheet_name}_{key}"
                    self._process_and_write_sheet(wb, value, new_sheet_name)
                    
                    cell.value = f"[Xem sheet: {new_sheet_name[:31]}] ({len(value)} mục)"
                    cell.hyperlink = f"#'{new_sheet_name[:31]}'!A1"
                    cell.font = hyperlink_font
                else:
                    if isinstance(value, (list, tuple)):
                        cell.value = str(value)
                    else:
                        cell.value = value

    def extract_file(self, data_to_extract, name_algorithm, filename="export_data.xlsx"):
        try:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            data_source = data_to_extract if isinstance(data_to_extract, list) else [data_to_extract]
            
            if not data_source:
                print("Lỗi: Dữ liệu rỗng.")
                return

            self._process_and_write_sheet(wb, data_source, "Main_Data")

            # Lưu file
            output_filename = f"{name_algorithm}_{filename}"
            wb.save(output_filename)

        except Exception as e:
            print(f"Đã xảy ra lỗi khi xuất file: {e}")

    def effect_movement(self):
        self.screen.blit(self.image, self.rect)
    
    def check_collided(self, pos):
        return self.rect.collidepoint(pos)

